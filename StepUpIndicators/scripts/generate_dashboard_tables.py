#!/usr/bin/env python3
from __future__ import annotations

import html
import math
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


@dataclass
class PageConfig:
    source_xlsx: str
    output_html: str
    page_title: str
    category_label: str
    description: str
    header_row: int
    data_start_row: int
    back_href: str = "impact-categories.html"
    back_text: str = "Back to Impact Categories"
    header_overrides: Optional[Dict[int, str]] = None


PAGES: List[PageConfig] = [
    PageConfig(
        source_xlsx="Dashboard/1_Economics.xlsx",
        output_html="economics.html",
        page_title="Economics",
        category_label="Category 01",
        description="Browse economic indicators, factors, and methodologies used across livestock systems.",
        header_row=1,
        data_start_row=2,
    ),
    PageConfig(
        source_xlsx="Dashboard/2_TCA.xlsx",
        output_html="tca.html",
        page_title="True Cost Accounting (TCA)",
        category_label="Category 02",
        description="Explore True Cost Accounting indicators and valuation methods across livestock systems.",
        header_row=2,
        data_start_row=3,
    ),
    PageConfig(
        source_xlsx="Dashboard/3_sLCA.xlsx",
        output_html="social.html",
        page_title="Social Life Cycle Assesments (s-LCA)",
        category_label="Category 03",
        description="Explore social life cycle assessment indicators and impact subcategories across production systems.",
        header_row=1,
        data_start_row=2,
    ),
    PageConfig(
        source_xlsx="Dashboard/4_enteric.xlsx",
        output_html="enteric.html",
        page_title="Greenhouse Gases: Enteric",
        category_label="Category 04",
        description="Review enteric methane indicators, equations, and evidence from the literature.",
        header_row=1,
        data_start_row=2,
        back_href="greenhouse-gases.html",
        back_text="Back to Greenhouse Gases",
        header_overrides={0: "Nr"},
    ),
    PageConfig(
        source_xlsx="Dashboard/4_Manure.xlsx",
        output_html="manure.html",
        page_title="Greenhouse Gases: Manure",
        category_label="Category 04",
        description="Review manure-related greenhouse gas indicators, methods, and results.",
        header_row=1,
        data_start_row=2,
        back_href="greenhouse-gases.html",
        back_text="Back to Greenhouse Gases",
    ),
    PageConfig(
        source_xlsx="Dashboard/5_AHW.xlsx",
        output_html="AHW.html",
        page_title="Animal Health and Welfare",
        category_label="Category 05",
        description="Browse animal health and welfare indicators, classifications, and applications.",
        header_row=1,
        data_start_row=2,
    ),
    PageConfig(
        source_xlsx="Dashboard/6_soilhealth.xlsx",
        output_html="soil-health.html",
        page_title="Soil Health",
        category_label="Category 06",
        description="Explore soil health indicators and evidence linked to livestock management systems.",
        header_row=1,
        data_start_row=2,
    ),
    PageConfig(
        source_xlsx="Dashboard/7_Biodiversity.xlsx",
        output_html="biodiversity.html",
        page_title="Biodiversity",
        category_label="Category 07",
        description="Explore biodiversity indicators and outcomes associated with livestock systems.",
        header_row=1,
        data_start_row=2,
    ),
    PageConfig(
        source_xlsx="Dashboard/8_nLCA.xlsx",
        output_html="n-lca.html",
        page_title="Nutritional Life Cycle Assessment (n-LCA)",
        category_label="Category 08",
        description="Review nutritional life cycle assessment indicators and integrated outcomes.",
        header_row=1,
        data_start_row=2,
    ),
]


PAGE_TEMPLATE = """<!DOCTYPE html>
<html lang=\"en\">
<head>
    <meta charset=\"UTF-8\">
    <meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
    <title>{title}</title>
    <link href=\"https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css\" rel=\"stylesheet\">
    <link href=\"https://cdn.datatables.net/1.13.6/css/dataTables.bootstrap5.min.css\" rel=\"stylesheet\">
    <link href=\"https://cdn.datatables.net/buttons/2.4.1/css/buttons.bootstrap5.min.css\" rel=\"stylesheet\">
    <style>
        :root {{
            --amarillo: #DDC646;
            --verde1: #72AE4C;
            --verde2: #305D3B;
            --celeste1: #E4F8FC;
            --celeste2: #9BD7E0;
        }}

        body {{
            background-color: #ffffff;
            font-family: 'Georgia', serif;
            color: #333;
            line-height: 1.6;
            padding: 0;
            margin: 0;
        }}

        .container-fluid {{
            max-width: 1600px;
            padding: 60px 40px;
        }}

        .navigation {{
            margin-bottom: 30px;
            padding-bottom: 20px;
            border-bottom: 1px solid #e0e0e0;
        }}

        .back-link {{
            font-size: 0.95rem;
            color: var(--verde2);
            text-decoration: none;
            display: inline-flex;
            align-items: center;
            transition: color 0.2s ease;
        }}

        .back-link:hover {{
            color: var(--verde1);
        }}

        .back-arrow {{
            margin-right: 8px;
            font-size: 1.2rem;
        }}

        .header {{
            margin-bottom: 40px;
            padding-bottom: 30px;
            border-bottom: 4px solid var(--verde2);
        }}

        .category-label {{
            font-size: 0.85rem;
            text-transform: uppercase;
            letter-spacing: 1.5px;
            color: #999;
            margin-bottom: 10px;
        }}

        h1 {{
            font-size: 3rem;
            font-weight: 700;
            color: var(--verde2);
            margin: 0;
            letter-spacing: -0.5px;
            line-height: 1.1;
        }}

        .description {{
            font-size: 1.1rem;
            color: #666;
            margin-top: 15px;
            line-height: 1.7;
        }}

        .table-container {{
            background: white;
            margin-top: 40px;
        }}

        .table-header-section {{
            margin-bottom: 25px;
            padding-bottom: 15px;
            border-bottom: 1px solid #e0e0e0;
        }}

        .table-title {{
            font-size: 1.4rem;
            font-weight: 700;
            color: var(--verde2);
            margin: 0;
        }}

        .table-wrapper {{
            overflow-x: auto;
        }}

        table.dataTable {{
            border-collapse: collapse;
            width: 100%;
            border: none;
            font-size: 0.9rem;
        }}

        table.dataTable thead th {{
            background-color: var(--verde2);
            color: white;
            border: none;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            padding: 14px 10px;
            font-size: 0.74rem;
            vertical-align: middle;
            white-space: nowrap;
            text-align: center;
        }}

        table.dataTable tbody td {{
            border-bottom: 1px solid #e0e0e0;
            padding: 12px 10px;
            vertical-align: top;
            line-height: 1.5;
            text-align: center;
            min-width: 110px;
            white-space: normal;
            word-break: break-word;
        }}

        table.dataTable tbody tr:hover {{
            background-color: #f8f8f8;
        }}

        .dt-buttons {{
            margin-bottom: 15px;
        }}

        .dt-buttons .btn {{
            background-color: white;
            border: 1px solid #ddd;
            color: var(--verde2);
            padding: 8px 14px;
            font-size: 0.84rem;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            border-radius: 0;
            margin-right: 5px;
            transition: all 0.2s ease;
        }}

        .dt-buttons .btn:hover {{
            background-color: var(--verde2);
            border-color: var(--verde2);
            color: white;
        }}

        .dataTables_wrapper .dataTables_paginate .paginate_button.current {{
            background: var(--verde2) !important;
            border-color: var(--verde2) !important;
            color: white !important;
        }}

        @media (max-width: 768px) {{
            h1 {{
                font-size: 2rem;
            }}

            .container-fluid {{
                padding: 30px 20px;
            }}
        }}
    </style>
</head>
<body>
    <div class=\"container-fluid\">
        <div class=\"navigation\">
            <a href=\"{back_href}\" class=\"back-link\">
                <span class=\"back-arrow\">←</span> {back_text}
            </a>
        </div>

        <div class=\"header\">
            <div class=\"category-label\">{category_label}</div>
            <h1>{title}</h1>
            <p class=\"description\">{description}</p>
        </div>

        <div class=\"table-container\">
            <div class=\"table-header-section\">
                <h2 class=\"table-title\">Research Database</h2>
            </div>
            <div class=\"table-wrapper\">
                <table id=\"dataTable\" class=\"table table-hover\" style=\"width:100%\">
                    <thead>
                        <tr>
{thead}
                        </tr>
                    </thead>
                    <tbody>
{tbody}
                    </tbody>
                </table>
            </div>
        </div>
    </div>

    <script src=\"https://code.jquery.com/jquery-3.7.0.min.js\"></script>
    <script src=\"https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js\"></script>
    <script src=\"https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js\"></script>
    <script src=\"https://cdn.datatables.net/1.13.6/js/dataTables.bootstrap5.min.js\"></script>
    <script src=\"https://cdn.datatables.net/buttons/2.4.1/js/dataTables.buttons.min.js\"></script>
    <script src=\"https://cdn.datatables.net/buttons/2.4.1/js/buttons.bootstrap5.min.js\"></script>
    <script src=\"https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js\"></script>
    <script src=\"https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.2.7/pdfmake.min.js\"></script>
    <script src=\"https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.2.7/vfs_fonts.js\"></script>
    <script src=\"https://cdn.datatables.net/buttons/2.4.1/js/buttons.html5.min.js\"></script>
    <script src=\"https://cdn.datatables.net/buttons/2.4.1/js/buttons.print.min.js\"></script>

    <script>
        $(document).ready(function() {{
            $('#dataTable').DataTable({{
                pageLength: 10,
                lengthMenu: [[5, 10, 25, 50, -1], [5, 10, 25, 50, 'All']],
                dom: 'Blfrtip',
                buttons: ['copy', 'csv', 'excel', 'pdf', 'print'],
                language: {{
                    search: 'Search:',
                    lengthMenu: 'Show _MENU_ entries',
                    info: 'Showing _START_ to _END_ of _TOTAL_ entries',
                    infoEmpty: 'Showing 0 to 0 of 0 entries',
                    infoFiltered: '(filtered from _MAX_ total entries)',
                    paginate: {{
                        first: 'First',
                        last: 'Last',
                        next: 'Next',
                        previous: 'Previous'
                    }}
                }},
                responsive: false,
                scrollX: true,
                autoWidth: false,
                order: []
            }});
        }});
    </script>
</body>
</html>
"""


def _clean_text(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, (datetime, date, time)):
        text = value.isoformat(sep=" ") if isinstance(value, datetime) else value.isoformat()
    elif isinstance(value, float):
        if math.isnan(value):
            return ""
        text = str(int(value)) if value.is_integer() else f"{value:.15g}"
    else:
        text = str(value)

    text = text.replace("\r\n", "\n").replace("\r", "\n").strip()
    text = "".join(ch for ch in text if ch == "\n" or ch == "\t" or ord(ch) >= 32)
    return text


def _html_escape(value: str) -> str:
    return html.escape(value, quote=True).replace("\n", "<br>")


def _load_table(cfg: PageConfig) -> tuple[list[str], list[list[str]]]:
    workbook = load_workbook(ROOT / cfg.source_xlsx, data_only=True, read_only=True)
    sheet = workbook[workbook.sheetnames[0]]

    all_rows = list(
        sheet.iter_rows(
            min_row=cfg.header_row,
            max_row=sheet.max_row,
            max_col=sheet.max_column,
            values_only=True,
        )
    )
    workbook.close()

    if not all_rows:
        raise ValueError(f"No rows found in {cfg.source_xlsx}")

    header_values = list(all_rows[0])
    data_offset = cfg.data_start_row - cfg.header_row
    data_values = [list(r) for r in all_rows[data_offset:]]

    used_cols: list[int] = []
    for col_idx in range(len(header_values)):
        header_cell = _clean_text(header_values[col_idx])
        has_override = bool(cfg.header_overrides and col_idx in cfg.header_overrides)
        if header_cell or has_override:
            used_cols.append(col_idx)

    headers: list[str] = []
    for out_idx, col_idx in enumerate(used_cols):
        if cfg.header_overrides and col_idx in cfg.header_overrides:
            header_text = cfg.header_overrides[col_idx]
        else:
            header_text = _clean_text(header_values[col_idx])

        headers.append(header_text if header_text else f"Column {out_idx + 1}")

    rows: list[list[str]] = []
    for row in data_values:
        cells = [_clean_text(row[col_idx] if col_idx < len(row) else None) for col_idx in used_cols]
        if any(cells):
            rows.append(cells)

    return headers, rows


def _build_html(cfg: PageConfig, headers: list[str], rows: list[list[str]]) -> str:
    thead = "\n".join(f"                            <th>{_html_escape(col)}</th>" for col in headers)

    body_rows: list[str] = []
    for row in rows:
        tds = "\n".join(f"                                <td>{_html_escape(value)}</td>" for value in row)
        body_rows.append("                        <tr>\n" + tds + "\n                        </tr>")

    tbody = "\n".join(body_rows)

    return PAGE_TEMPLATE.format(
        title=_html_escape(cfg.page_title),
        category_label=_html_escape(cfg.category_label),
        description=_html_escape(cfg.description),
        back_href=_html_escape(cfg.back_href),
        back_text=_html_escape(cfg.back_text),
        thead=thead,
        tbody=tbody,
    )


def main() -> None:
    for cfg in PAGES:
        headers, rows = _load_table(cfg)
        html_doc = _build_html(cfg, headers, rows)
        out_path = ROOT / cfg.output_html
        out_path.write_text(html_doc, encoding="utf-8")
        print(f"Generated {cfg.output_html}: {len(rows)} rows, {len(headers)} columns")


if __name__ == "__main__":
    main()
